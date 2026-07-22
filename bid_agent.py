#!/usr/bin/env python3
"""零碳园区招标信息聚合 Agent — 自动化搜索、大模型抽取、全量验证与增量持久化"""

import argparse
import asyncio
import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import aiohttp
from dotenv import load_dotenv
from openai import (
    APIConnectionError,
    APIError,
    AsyncOpenAI,
    InternalServerError,
    RateLimitError,
)
from openai.types.chat import ChatCompletion
from pydantic import BaseModel, Field
from tavily import AsyncTavilyClient
from tenacity import (
    retry,
    retry_if_exception_type,
    stop_after_attempt,
    wait_exponential,
)

# Local schema
from schemas import FIELD_META

load_dotenv()

# ---------------------------------------------------------------------------
# Paths & Constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
JSONL_FILE = BASE_DIR / "evaluated_bids.jsonl"
OUTPUT_FILE = BASE_DIR / "bids_data.json"
CONCURRENCY = 5

TAVILY_KEY = os.getenv("TAVILY_API_KEY", "")
DEEPSEEK_KEY = os.getenv("DEEPSEEK_API_KEY", "")
SERPER_KEY = os.getenv("SERPER_API_KEY", "")
TALORDATA_KEY = os.getenv("TALORDATA_API_KEY", "")
TALORDATA_URL = "https://serpapi.talordata.net/serp/v1/request"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("bid_agent")


# ---------------------------------------------------------------------------
# Data Models
# ---------------------------------------------------------------------------

class BidRecord(BaseModel):
    """零碳园区招标信息完整记录 — 25 字段"""
    id: int = 0
    title: str = ""
    source_url: str = ""
    source_platform: str = ""
    crawled_at: str = ""
    province: str = ""
    city: str | None = None
    district: str | None = None
    bid_status: str = ""
    content_tags: list[str] = Field(default_factory=list)
    project_stage: str = ""
    publish_date: str | None = None
    deadline_date: str | None = None
    award_date: str | None = None
    amount_raw: str = ""
    amount_range: str = ""
    bidder: str = ""
    winner: str | None = None
    winner_type: str | None = None
    contact_info: str | None = None
    verified: bool = False
    verification_note: str = ""
    data_quality: str = ""
    park_level: str | None = None
    raw_snippets: str = ""
    remarks: str = ""


# ---------------------------------------------------------------------------
# System Prompt — 招标信息抽取
# ---------------------------------------------------------------------------

EXTRACTION_SYSTEM = """你是一个专业的政府采购与招标信息分析助手。请从以下搜索结果中，提取所有与"零碳园区"相关的招标/采购/中标公告的结构化信息。

## 提取规则

1. **仅提取明确提及零碳园区、近零碳园区、碳中和园区、低碳工业园等关键词的招标信息**
2. 如果搜索结果中某条不是招标/采购/中标相关公告，跳过它
3. 每条招标信息输出一个 JSON 对象
4. 所有字段必须从原文中提取，**严禁编造**
5. **每条 snippet 开头标注了 `URL: ...`，请将该 URL 填入 source_url 字段**

## 字段说明

- `title`: 公告标题，原文照录
- `source_url`: 公告链接
- `source_platform`: 来源平台（如"中国政府采购网""XX省公共资源交易平台""XX园区官网"）
- `province`: 省/自治区/直辖市（从公告内容推断）
- `city`: 地级市，如无法确定则为 null
- `district`: 区/县，如无法确定则为 null
- `bid_status`: 招标状态，必须是以下之一：
  - "预告/公告" — 采购意向、招标预告、资格预审
  - "招标中(含更正)" — 正式招标公告、更正公告、答疑澄清
  - "中标公示" — 中标/成交结果公示、合同公告
- `content_tags`: 工程内容标签（可多选），从以下列表中选择涉及的：
  ["可再生能源", "储能与微电网", "节能改造", "数字化碳管理", "绿色建筑与基础设施", "水处理与循环利用", "CCUS", "规划咨询与认证", "综合类"]
- `project_stage`: 项目阶段，必须是以下之一：
  "规划设计" / "工程建设" / "设备采购" / "运营服务" / "咨询认证"
- `publish_date`: 发布日期，格式 YYYY-MM-DD，如无则为 null
- `deadline_date`: 投标截止日期，格式 YYYY-MM-DD，如无则为 null
- `award_date`: 中标日期，格式 YYYY-MM-DD，如无则为 null
- `amount_raw`: 原始金额表述，原文照搬（如"预算金额500万元""中标金额1,234.56万元"），如无则为 ""
- `amount_range`: 金额区间，必须是以下之一：
  "<100万" / "100-500万" / "500-1000万" / "1000万-5000万" / "5000万-1亿" / ">1亿" / "未披露"
  注意：根据 amount_raw 提取的数值归类到对应区间，万元为单位的"500万"归入"100-500万"
- `bidder`: 招标方/采购人/业主名称
- `winner`: 中标方/成交供应商名称，如无则为 null
- `winner_type`: 中标方类型（"央企"/"国企"/"民企"/"外资"/"联合体"），如无则为 null
- `contact_info`: 招标公告中的联系人、电话或地址，原文照录，如无则为 null
- `data_quality`: 数据完整度，必须是以下之一：
  - "完整" — 标题、金额、日期、参与方均齐全
  - "部分缺失" — 缺少部分字段但核心信息（标题+日期+招标方）齐全
  - "仅标题" — 只有标题和链接，其他信息极少

## 输出格式

必须只输出一个 JSON 对象（不是数组），不要输出任何解释、分析或 Markdown 标记：
```json
{
  "bids": [
    {
      "title": "...",
      "source_url": "...",
      "source_platform": "...",
      "province": "...",
      "city": "..."或null,
      "district": "..."或null,
      "bid_status": "...",
      "content_tags": ["...", "..."],
      "project_stage": "...",
      "publish_date": "..."或null,
      "deadline_date": "..."或null,
      "award_date": "..."或null,
      "amount_raw": "...",
      "amount_range": "...",
      "bidder": "...",
      "winner": "..."或null,
      "winner_type": "..."或null,
      "contact_info": "..."或null,
      "data_quality": "..."
    }
  ]
}
```

如果没有找到任何零碳园区招标信息，输出 `{"bids": []}`。
只输出 JSON 对象，不要输出任何其他文字。"""


# ---------------------------------------------------------------------------
# Verification Prompt
# ---------------------------------------------------------------------------

VERIFICATION_SYSTEM = """你是一个严格的数据审核员。你的任务是逐字段核对一条招标信息记录是否与原始网页内容一致。

## 验证规则

1. 检查每个关键字段（title, amount_raw, bidder, winner, publish_date, deadline_date, bid_status）是否能在原文中找到支撑
2. 标题相似度需 > 80%（允许少量措辞差异）
3. 金额数字必须在原文中出现（忽略"万元""亿元"等单位表述差异）
4. 日期（年-月-日）必须在原文中出现
5. 招标方/中标方名称必须在原文中出现

## 输出格式

只输出一个 JSON 对象：
```json
{
  "verified": true/false,
  "verification_note": "如果 verified=true，写'全部关键字段有原文支撑'；如果 verified=false，逐条列出不匹配的具体字段及原文中实际内容"
}
```

只输出 JSON，不要输出任何其他文字。"""


# ---------------------------------------------------------------------------
# Persistence Helpers
# ---------------------------------------------------------------------------

def load_completed_urls() -> set[str]:
    """读取 jsonl 中已有的 source_url，实现去重"""
    if not JSONL_FILE.exists():
        return set()
    urls: set[str] = set()
    with open(JSONL_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                if obj.get("source_url"):
                    urls.add(obj["source_url"])
            except json.JSONDecodeError:
                continue
    return urls


def get_next_id() -> int:
    """获取下一个自增 ID"""
    if not JSONL_FILE.exists():
        return 1
    max_id = 0
    with open(JSONL_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                max_id = max(max_id, obj.get("id", 0))
            except json.JSONDecodeError:
                continue
    return max_id + 1


def append_record(rec: BidRecord) -> None:
    """追加一行 JSON 到 jsonl 文件"""
    rec.crawled_at = datetime.now(timezone.utc).isoformat()
    with open(JSONL_FILE, "a", encoding="utf-8") as f:
        f.write(rec.model_dump_json() + "\n")
    logger.info("  ✓ 已写入 jsonl: id=%d, title=%s", rec.id, rec.title[:50])


def jsonl_to_json() -> None:
    """将所有 jsonl 记录汇总导出为 bids_data.json"""
    if not JSONL_FILE.exists():
        logger.warning("jsonl 文件不存在，跳过导出")
        return
    records: list[dict] = []
    with open(JSONL_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    records.sort(key=lambda r: r["id"])
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    logger.info("导出 %d 条记录 → %s", len(records), OUTPUT_FILE)


# ---------------------------------------------------------------------------
# Search
# ---------------------------------------------------------------------------

SERPER_URL = "https://google.serper.dev/search"

# 31 省级行政区
PROVINCES = [
    "北京", "天津", "河北", "山西", "内蒙古",
    "辽宁", "吉林", "黑龙江",
    "上海", "江苏", "浙江", "安徽", "福建", "江西", "山东",
    "河南", "湖北", "湖南", "广东", "广西", "海南",
    "重庆", "四川", "贵州", "云南", "西藏",
    "陕西", "甘肃", "青海", "宁夏", "新疆",
]


def _build_search_queries() -> list[str]:
    """动态生成搜索关键词：基础 + 省份轮询 + 追踪"""
    base = [
        '"零碳园区" 招标公告',
        '"零碳产业园" 招标',
        '"零碳园区" 中标',
        '"零碳产业园" 中标公告',
        '"近零碳园区" 招标',
        '"碳中和园区" 采购',
        '"零碳园区" 采购公告',
        'site:gov.cn "零碳园区" 招标',
        'site:ccgp.gov.cn "零碳" "园区"',
        '"零碳园区" 竞争性磋商',
        '"零碳园区" EPC',
        '"零碳园区" 规划设计',
        # 追踪性关键词
        '"零碳园区" 成交公告',
        '"零碳园区" 合同公告',
        '"零碳园区" 中标候选人',
    ]
    province_queries = []
    for p in PROVINCES:
        province_queries.append(f'"{p}" "零碳园区" 招标')
        province_queries.append(f'"{p}" "零碳园区" 中标')        # 工程内容细化搜索（补漏）
    content_queries = [
        '"零碳园区" 光伏 EPC',
        '"零碳园区" 储能项目 中标',
        '"零碳园区" 微电网',
        '"零碳园区" 充电桩',
        '"零碳园区" 绿色建筑 改造',
        '"零碳园区" 能碳管理平台',
        '"零碳园区" 智慧能源管理系统',
        '"零碳园区" 源网荷储',
        '"零碳园区" 分布式光伏',
        '"零碳园区" 污水处理 招标',
        '"零碳园区" 余热回收',
        '"零碳园区" 风电',
        '"零碳园区" 中标候选人',
        '"零碳园区" 成交结果公告',
        '"零碳园区" 合同公告',
        '"零碳园区" 询价采购',
        '"零碳园区" 单一来源',
        '"零碳园区" 竞争性谈判',
        '"近零碳园区" 中标 公告',
        '"低碳产业园" "零碳" 招标',
        # 补强弱势类别
        '"零碳园区" CCUS 碳捕集',
        '"零碳园区" 氢能',
        '"零碳园区" 光储充',
        '"零碳园区" 虚拟电厂',
        '"零碳园区" LEED 绿标',
        '"零碳园区" 碳核查',
        '"零碳园区" 全过程咨询',
        '"零碳园区" 创建 服务 采购',
        '"零碳园区" 申报 服务 招标',
        '"零碳" "产业园" EPC 总承包',
        '"西藏" "零碳" 招标',
        '"新疆" "零碳园区" 项目 招标',
        '"云南" "零碳" 园区 招标',
    ]
    return base + province_queries + content_queries + __import__("json").loads(open(__import__("pathlib").Path(__file__).parent / "park_queries.json", "r", encoding="utf-8").read())


# 52个国家级园区定向搜索

SEARCH_QUERIES = _build_search_queries()


async def _search_serper(
    session: aiohttp.ClientSession, query: str,
) -> list[dict]:
    """通过 Serper (Google Search) API 搜索中文内容"""
    headers = {"X-API-KEY": SERPER_KEY, "Content-Type": "application/json"}
    payload = {"q": query, "gl": "cn", "hl": "zh-cn", "num": 5}
    try:
        async with session.post(SERPER_URL, json=payload, headers=headers,
                                timeout=aiohttp.ClientTimeout(total=15)) as resp:
            if resp.status != 200:
                logger.warning("  Serper 返回 %d: %s", resp.status, (await resp.text())[:100])
                return []
            data = await resp.json()
            return data.get("organic", [])
    except Exception as e:
        logger.warning("  Serper 搜索异常: %s", e)
        return []


async def _search_talordata(
    session: aiohttp.ClientSession, query: str,
) -> list[dict]:
    """通过 TalorData SERP API 搜索中文内容"""
    if not TALORDATA_KEY:
        return []
    headers = {
        "Authorization": f"Bearer {TALORDATA_KEY}",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    data = {
        "engine": "google",
        "q": query,
        "gl": "cn",
        "hl": "zh-CN",
        "json": "1",
    }
    try:
        async with session.post(TALORDATA_URL, headers=headers, data=data,
                                timeout=aiohttp.ClientTimeout(total=15)) as resp:
            if resp.status != 200:
                return []
            result = await resp.json()
            if result.get("code") != 0:
                return []
            return result.get("data", {}).get("organic", [])
    except Exception as e:
        logger.warning("  TalorData 搜索异常: %s", e)
        return []


async def search_all(
    tavily: AsyncTavilyClient,
    session: aiohttp.ClientSession | None,
    seen_urls: set[str],
) -> tuple[list[str], str, dict[str, str]]:
    """多关键词并行搜索，合并去重。
    返回: (all_urls, combined_snippets, url_to_snippet_map)
    """
    all_urls: list[str] = []
    all_snippets: list[str] = []
    url_snippet_map: dict[str, str] = {}

    async def _tavily_search(query: str, label: str):
        try:
            resp = await tavily.search(
                query, search_depth="advanced", max_results=5
            )
            for r in resp.get("results", []):
                url = r.get("url", "")
                content = r.get("content") or r.get("snippet", "")
                if url and url not in seen_urls and url not in all_urls:
                    all_urls.append(url)
                if content:
                    snippet_text = f"[{label}] {content[:1500]}"
                    all_snippets.append(f"[{label}] URL: {url}\n{content[:2000]}")
                    # 保存 per-url snippet（取最长的）
                    if url not in url_snippet_map or len(snippet_text) > len(url_snippet_map[url]):
                        url_snippet_map[url] = snippet_text
            logger.debug("  Tavily [%s] '%s' → %d results",
                         label, query[:30], len(resp.get("results", [])))
        except Exception:
            logger.warning("  Tavily [%s] 搜索失败: %s", label, query[:40])

    async def _serper_search(query: str, label: str):
        if not SERPER_KEY or session is None:
            return
        try:
            results = await _search_serper(session, query)
            for r in results:
                url = r.get("link", "")
                snippet = r.get("snippet", "")
                if url and url not in seen_urls and url not in all_urls:
                    all_urls.append(url)
                if snippet:
                    snippet_text = f"[{label}] {snippet[:1500]}"
                    all_snippets.append(f"[{label}] URL: {url}\n{snippet[:2000]}")
                    if url not in url_snippet_map or len(snippet_text) > len(url_snippet_map[url]):
                        url_snippet_map[url] = snippet_text
            logger.debug("  Serper [%s] '%s' → %d results",
                         label, query[:30], len(results))
        except Exception:
            logger.warning("  Serper [%s] 搜索失败: %s", label, query[:40])

    async def _talordata_search(query: str, label: str):
        if not TALORDATA_KEY or session is None:
            return
        try:
            results = await _search_talordata(session, query)
            for r in results:
                url = r.get("link", "")
                snippet = r.get("snippet", "") or r.get("description", "")
                if url and url not in seen_urls and url not in all_urls:
                    all_urls.append(url)
                if snippet:
                    snippet_text = f"[{label}] {snippet[:1500]}"
                    all_snippets.append(f"[{label}] URL: {url}\n{snippet[:2000]}")
                    if url not in url_snippet_map or len(snippet_text) > len(url_snippet_map[url]):
                        url_snippet_map[url] = snippet_text
            logger.debug("  TalorData [%s] '%s' -> %d results",
                         label, query[:30], len(results))
        except Exception:
            logger.warning("  TalorData [%s] 搜索失败: %s", label, query[:40])

    # 并行执行所有搜索（Tavily + Serper + TalorData 三通道）
    tasks = []
    for i, q in enumerate(SEARCH_QUERIES):
        tasks.append(_tavily_search(q, f"TV{i}"))
        tasks.append(_serper_search(q, f"SP{i}"))
        tasks.append(_talordata_search(q, f"TD{i}"))
    await asyncio.gather(*tasks)

    combined = "\n---\n".join(all_snippets)
    logger.info("搜索完成: %d 条新 URL, %d 字符文本", len(all_urls), len(combined))
    return all_urls, combined, url_snippet_map


# ---------------------------------------------------------------------------
# LLM Calls
# ---------------------------------------------------------------------------

async def _call_llm(
    aclient: AsyncOpenAI, model: str, messages: list[dict],
    max_tokens: int = 2000,
) -> ChatCompletion:
    """底层 LLM 调用"""
    return await aclient.chat.completions.create(
        model=model,
        messages=messages,
        max_tokens=max_tokens,
        temperature=0.3,
        response_format={"type": "json_object"},
    )


# tenacity 重试包装
_call_llm_retry = retry(
    retry=retry_if_exception_type((
        OSError, ConnectionError, TimeoutError,
        APIConnectionError, RateLimitError, InternalServerError,
    )),
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=30),
)(_call_llm)


async def extract_bids(
    aclient: AsyncOpenAI, model: str, snippets: str,
) -> list[dict]:
    """调用 DeepSeek 从搜索片段中提取招标信息"""
    messages = [
        {"role": "system", "content": EXTRACTION_SYSTEM},
        {"role": "user", "content": f"请从以下搜索结果中提取所有零碳园区相关的招标/采购信息：\n\n{snippets[:15000]}"},
    ]
    try:
        resp = await _call_llm_retry(aclient, model, messages, max_tokens=12000)
        raw = resp.choices[0].message.content or "{}"
        # 清理可能的 markdown 代码块
        raw = raw.strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
        data = json.loads(raw)
        # 标准格式: {"bids": [...]}
        if isinstance(data, dict):
            return data.get("bids", [])
        return []
    except json.JSONDecodeError as e:
        logger.warning("  JSON 解析失败: %s, raw=%s...", e, (raw or "")[:300])
        return []
    except (APIError, APIConnectionError, RateLimitError, InternalServerError) as e:
        logger.error("  LLM 调用失败: %s", e)
        raise


async def verify_record(
    aclient: AsyncOpenAI, model: str, record: dict, page_text: str,
) -> tuple[bool, str]:
    """对比原始网页内容验证提取的字段"""
    # 准备简洁的验证输入
    fields_to_check = {
        "title": record.get("title", ""),
        "amount_raw": record.get("amount_raw", ""),
        "bidder": record.get("bidder", ""),
        "winner": record.get("winner", ""),
        "publish_date": record.get("publish_date", ""),
        "deadline_date": record.get("deadline_date", ""),
        "bid_status": record.get("bid_status", ""),
    }
    check_json = json.dumps(fields_to_check, ensure_ascii=False, indent=2)

    messages = [
        {"role": "system", "content": VERIFICATION_SYSTEM},
        {"role": "user", "content": (
            f"## 提取的记录\n```json\n{check_json}\n```\n\n"
            f"## 原始网页内容\n{page_text[:8000]}"
        )},
    ]
    try:
        resp = await _call_llm_retry(aclient, model, messages, max_tokens=800)
        raw = resp.choices[0].message.content or "{}"
        raw = raw.strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
        result = json.loads(raw)
        return result.get("verified", False), result.get("verification_note", "")
    except (json.JSONDecodeError, KeyError) as e:
        logger.warning("  验证 JSON 解析失败: %s", e)
        return False, f"验证响应解析失败: {e}"
    except (APIError, APIConnectionError, RateLimitError, InternalServerError) as e:
        logger.error("  验证 LLM 调用失败: %s", e)
        raise


# ---------------------------------------------------------------------------
# Web Page Fetcher
# ---------------------------------------------------------------------------

async def fetch_page_text(
    session: aiohttp.ClientSession, url: str,
) -> str | None:
    """获取网页正文（HTML → 纯文本）"""
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/125.0.0.0 Safari/537.36"
        }
        async with session.get(url, headers=headers,
                               timeout=aiohttp.ClientTimeout(total=20)) as resp:
            if resp.status != 200:
                logger.debug("  URL 不可达 [%d]: %s", resp.status, url[:80])
                return None
            html = await resp.text(encoding="utf-8", errors="replace")
            # 简单 HTML → 文本
            text = _html_to_text(html)
            return text[:10000] if text else None
    except Exception as e:
        logger.debug("  URL 抓取异常: %s — %s", url[:80], e)
        return None


def _html_to_text(html: str) -> str:
    """简单 HTML 标签清理，提取纯文本"""
    # 移除 script/style
    text = re.sub(r'<(script|style)[^>]*>.*?</\1>', '', html, flags=re.DOTALL | re.IGNORECASE)
    # 移除 HTML 标签
    text = re.sub(r'<[^>]+>', ' ', text)
    # 解码常见 HTML 实体
    text = text.replace('&nbsp;', ' ').replace('&lt;', '<').replace('&gt;', '>')
    text = text.replace('&amp;', '&').replace('&quot;', '"').replace('&#39;', "'")
    text = text.replace('&mdash;', '—').replace('&ndash;', '–')
    # 合并空白
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n\s*\n', '\n', text)
    return text.strip()


# ---------------------------------------------------------------------------
# Main Pipeline
# ---------------------------------------------------------------------------

async def process_one(
    sem: asyncio.Semaphore,
    tavily: AsyncTavilyClient,
    session: aiohttp.ClientSession,
    aclient: AsyncOpenAI,
    model: str,
    next_id: int,
) -> list[BidRecord]:
    """单轮搜索 → 抽取 → 验证 → 持久化"""
    async with sem:
        seen_urls = load_completed_urls()
        logger.info("=" * 60)
        logger.info("开始新一轮搜索... 已入库 %d 条记录", len(seen_urls))

        # Step 1: 搜索
        urls, snippets, url_snippet_map = await search_all(tavily, session, seen_urls)
        if not urls:
            logger.info("本轮无新 URL，结束")
            return []
        logger.info("搜索到 %d 条新 URL, %d 字符文本片段", len(urls), len(snippets))

        # Step 2: 抽取
        raw_records = await extract_bids(aclient, model, snippets)
        if not raw_records:
            logger.info("LLM 未提取到任何招标信息")
            return []
        logger.info("LLM 提取到 %d 条候选记录", len(raw_records))

        # Step 3: 逐条处理
        results: list[BidRecord] = []
        current_id = next_id

        for i, raw in enumerate(raw_records):
            logger.info("--- 处理 %d/%d: %s", i + 1, len(raw_records),
                        raw.get("title", "?")[:60])

            # 去重检查
            source_url = raw.get("source_url", "")
            if source_url and source_url in seen_urls:
                logger.info("  跳过：URL 已存在")
                continue

            # 构建记录
            rec = BidRecord(
                id=current_id,
                title=raw.get("title") or "",
                source_url=raw.get("source_url") or "",
                source_platform=raw.get("source_platform") or "",
                province=raw.get("province") or "",
                city=raw.get("city"),
                district=raw.get("district"),
                bid_status=raw.get("bid_status") or "",
                content_tags=raw.get("content_tags") or [],
                project_stage=raw.get("project_stage") or "",
                publish_date=raw.get("publish_date"),
                deadline_date=raw.get("deadline_date"),
                award_date=raw.get("award_date"),
                amount_raw=raw.get("amount_raw") or "",
                amount_range=raw.get("amount_range") or "",
                bidder=raw.get("bidder") or "",
                winner=raw.get("winner"),
                winner_type=raw.get("winner_type"),
                contact_info=raw.get("contact_info"),
                data_quality=raw.get("data_quality") or "",
                raw_snippets=url_snippet_map.get(source_url, ""),
            )

            # Step 4: 全量验证 — 抓取原文对比
            if source_url:
                page_text = await fetch_page_text(session, source_url)
                if page_text:
                    verified, vnote = await verify_record(aclient, model,
                                                          raw, page_text)
                    rec.verified = verified
                    rec.verification_note = vnote
                    logger.info("  验证: %s — %s",
                                "✓ 通过" if verified else "✗ 未通过",
                                vnote[:80] if vnote else "")
                else:
                    rec.verified = False
                    rec.verification_note = "无法访问原文 URL，无法完成验证"
                    logger.info("  验证: ✗ URL 不可达")
            else:
                rec.verified = False
                rec.verification_note = "无 source_url，无法验证"
                logger.info("  验证: ✗ 无来源链接")

            # Step 5: 持久化
            append_record(rec)
            results.append(rec)
            seen_urls.add(source_url)
            current_id += 1

        return results


async def main_async(args):
    """主入口"""
    # 初始化客户端
    tavily = AsyncTavilyClient(api_key=TAVILY_KEY)
    aclient = AsyncOpenAI(
        api_key=DEEPSEEK_KEY,
        base_url="https://api.deepseek.com",
    )
    model = "deepseek-chat"

    sem = asyncio.Semaphore(CONCURRENCY)
    all_results: list[BidRecord] = []

    async with aiohttp.ClientSession() as session:
        next_id = get_next_id()

        if args.limit:
            # 限制模式：只跑一轮搜索+抽取
            seen_urls = load_completed_urls()
            logger.info("限制模式：最多 %d 条", args.limit)
            urls, snippets, url_snippet_map = await search_all(tavily, session, seen_urls)
            if not urls:
                logger.info("无搜索结果")
                return

            raw_records = await extract_bids(aclient, model, snippets)
            if not raw_records:
                logger.info("LLM 未提取到任何招标信息")
                return

            raw_records = raw_records[:args.limit]
            logger.info("提取到 %d 条记录（截取前 %d 条）", len(raw_records), args.limit)

            current_id = next_id
            for i, raw in enumerate(raw_records):
                source_url = raw.get("source_url", "")
                if source_url and source_url in seen_urls:
                    logger.info("跳过重复: %s", source_url[:80])
                    continue

                rec = BidRecord(
                    id=current_id,
                    title=raw.get("title") or "",
                    source_url=source_url or "",
                    source_platform=raw.get("source_platform") or "",
                    province=raw.get("province") or "",
                    city=raw.get("city"),
                    district=raw.get("district"),
                    bid_status=raw.get("bid_status") or "",
                    content_tags=raw.get("content_tags") or [],
                    project_stage=raw.get("project_stage") or "",
                    publish_date=raw.get("publish_date"),
                    deadline_date=raw.get("deadline_date"),
                    award_date=raw.get("award_date"),
                    amount_raw=raw.get("amount_raw") or "",
                    amount_range=raw.get("amount_range") or "",
                    bidder=raw.get("bidder") or "",
                    winner=raw.get("winner"),
                    winner_type=raw.get("winner_type"),
                    contact_info=raw.get("contact_info"),
                    data_quality=raw.get("data_quality") or "",
                    raw_snippets=url_snippet_map.get(source_url, ""),
                )

                # 验证
                if source_url:
                    page_text = await fetch_page_text(session, source_url)
                    if page_text:
                        verified, vnote = await verify_record(aclient, model,
                                                              raw, page_text)
                        rec.verified = verified
                        rec.verification_note = vnote
                        logger.info("  [%d/%d] 验证: %s", i + 1, len(raw_records),
                                    "✓" if verified else "✗")
                    else:
                        rec.verified = False
                        rec.verification_note = "无法访问原文 URL"
                else:
                    rec.verified = False
                    rec.verification_note = "无 source_url"

                append_record(rec)
                all_results.append(rec)
                seen_urls.add(source_url)
                current_id += 1
        else:
            # 全量模式：持续搜索直到无新结果或达到 max_rounds
            max_rounds = args.rounds or 3
            for round_num in range(1, max_rounds + 1):
                logger.info("\n" + "=" * 60)
                logger.info("第 %d/%d 轮", round_num, max_rounds)
                results = await process_one(sem, tavily, session, aclient,
                                            model, get_next_id())
                if not results:
                    logger.info("第 %d 轮无新结果，提前结束", round_num)
                    break
                all_results.extend(results)

    # 最终导出
    jsonl_to_json()
    logger.info("=" * 60)
    logger.info("完成！共 %d 条招标记录 → %s", len(all_results), OUTPUT_FILE)

    # 生成 Excel
    if all_results or JSONL_FILE.exists():
        export_to_xlsx()


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def export_to_xlsx():
    """从 jsonl 导出 Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl 未安装，跳过 Excel 生成。安装: pip install openpyxl")
        return

    if not JSONL_FILE.exists():
        logger.warning("jsonl 文件不存在，跳过 Excel 导出")
        return

    records: list[dict] = []
    with open(JSONL_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    records.sort(key=lambda r: r["id"])

    wb = Workbook()
    ws = wb.active
    ws.title = "招标信息"

    # 样式
    HEADER_FILL = PatternFill(start_color="1A3A2A", end_color="1A3A2A", fill_type="solid")
    HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    CELL_FONT = Font(name="Microsoft YaHei", size=10)
    VERIFIED_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    FAILED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP = Alignment(vertical="top", wrap_text=True)

    # 表头
    headers = [m["header"] for m in FIELD_META]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 数据行
    for i, rec in enumerate(records):
        r = i + 2
        field_names = [m["field"] for m in FIELD_META]
        for col, fname in enumerate(field_names, 1):
            val = rec.get(fname, "")
            if isinstance(val, list):
                val = " / ".join(val)
            elif val is None:
                val = ""
            elif isinstance(val, bool):
                val = "是" if val else "否"

            cell = ws.cell(row=r, column=col, value=val)
            cell.font = CELL_FONT
            wrap_fields = ("title", "verification_note", "raw_snippets",
                           "remarks", "contact_info", "amount_raw")
            cell.alignment = WRAP if fname in wrap_fields else CENTER
            cell.border = THIN_BORDER

            # 验证状态着色
            if fname == "verified":
                if rec.get("verified"):
                    cell.fill = VERIFIED_FILL
                else:
                    cell.fill = FAILED_FILL

    # 列宽
    widths = {
        1: 5, 2: 40, 3: 40, 4: 18, 5: 18, 6: 8, 7: 8, 8: 8,
        9: 15, 10: 22, 11: 12, 12: 12, 13: 12, 14: 12,
        15: 22, 16: 14, 17: 24, 18: 24, 19: 12, 20: 28,
        21: 10, 22: 40, 23: 12, 24: 50, 25: 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Sheet 2: 分类说明
    ws2 = wb.create_sheet("分类说明")
    notes = [
        ["分类体系说明"],
        [""],
        ["一、招标状态"],
        ["预告/公告 — 采购意向公开、招标预告、资格预审公告"],
        ["招标中(含更正) — 正式招标公告、更正公告、答疑澄清文件"],
        ["中标公示 — 中标/成交结果公告、合同公告"],
        [""],
        ["二、工程内容（可多选）"],
        ["可再生能源 — 光伏、风电、生物质、地热等"],
        ["储能与微电网 — 电池储能、抽水蓄能、微电网、虚拟电厂"],
        ["节能改造 — 工业节能、建筑节能、余热回收、电机升级"],
        ["数字化碳管理 — 碳监测平台、数字孪生、EMS、碳核算"],
        ["绿色建筑与基础设施 — LEED/绿标、低碳交通、充电桩、海绵城市"],
        ["水处理与循环利用 — 污水处理、中水回用、废弃物资源化"],
        ["CCUS — 碳捕集、利用与封存"],
        ["规划咨询与认证 — 园区规划、碳核查、认证服务、可研报告"],
        ["综合类 — 同时覆盖多个领域的综合项目"],
        [""],
        ["三、项目阶段"],
        ["规划设计 / 工程建设 / 设备采购 / 运营服务 / 咨询认证"],
        [""],
        ["四、金额区间"],
        ["<100万 / 100-500万 / 500-1000万 / 1000万-5000万 / 5000万-1亿 / >1亿 / 未披露"],
    ]
    for i, row_data in enumerate(notes):
        cell = ws2.cell(row=i + 1, column=1, value=row_data[0])
        if i == 0:
            cell.font = Font(name="Microsoft YaHei", size=14, bold=True, color="1A3A2A")
        elif row_data[0].startswith(("一", "二", "三", "四")):
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
        else:
            cell.font = CELL_FONT
    ws2.column_dimensions["A"].width = 60

    xlsx_path = BASE_DIR / "bids_data.xlsx"
    wb.save(xlsx_path)
    logger.info("Excel 已生成: %s (%d 条记录)", xlsx_path, len(records))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="零碳园区招标信息聚合 Agent")
    parser.add_argument("--limit", type=int, default=0,
                        help="限制提取条数（测试用）")
    parser.add_argument("--rounds", type=int, default=5,
                        help="全量模式最大搜索轮数（默认 5）")
    args = parser.parse_args()

    if not TAVILY_KEY or not DEEPSEEK_KEY:
        logger.error("缺少 API Key！请在 .env 中配置 TAVILY_API_KEY 和 DEEPSEEK_API_KEY")
        sys.exit(1)

    asyncio.run(main_async(args))


if __name__ == "__main__":
    main()
